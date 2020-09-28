from openpyxl import load_workbook
from tkinter import messagebox
import getpass


def check_key(licence,frame,L1):
    user_name = getpass.getuser()
    print(type(user_name))
    path = "Licence.xlsx"
    book = load_workbook(filename=path, data_only=True)
    worksheet_names = book.sheetnames
    licenceplate = dict()
    for name in worksheet_names:
        worksheet = book[name]
        num_rows = worksheet.max_row
        curr_row = 1
        while curr_row < num_rows + 1:
            if worksheet.cell(curr_row, 2).value is not None:
                if worksheet.cell(curr_row, 3).value is not None:
                    licenceplate[worksheet.cell(curr_row, 2).value] = worksheet.cell(curr_row, 3).value
            curr_row += 1
        for l in licenceplate.keys():
            if l == int(licence):
                if licenceplate[l] == user_name:
                    L1.set("Hello word")
                    return True
                else:
                    messagebox.showerror(title="Wrong licence", message="Wrong computer with this software")
                    frame.update()
                    return False
        messagebox.showerror(title="License key error", message="There is an error of license")
        frame.update()
        return False
